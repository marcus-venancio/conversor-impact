import openpyxl

# Criar uma planilha(book)
book = openpyxl.Workbook()
# Como visualizar páginas existentes
print(book.sheetnames)
# Como criar uma página
book.create_sheet('Empresas')
# Como selecionar uma página
empresas_page = book ['Empresas']
empresas_page.append(['contato'])

#Salvar
book.save('Planilha com contato.xlsx')